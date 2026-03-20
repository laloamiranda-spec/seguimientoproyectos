from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import io
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import os
import sqlite3
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import date, datetime, timedelta
from dotenv import load_dotenv
load_dotenv(override=True)
# ── Neon (PostgreSQL) cuando DATABASE_URL está definida ─────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL')  # Neon provee esta variable
use_pg = bool(DATABASE_URL)

if use_pg:
    import psycopg2
    import psycopg2.extras

DB_PATH = 'entregables.db'  # Solo se usa si no hay DATABASE_URL

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'cambiame123')

# ── Flask-Login ──────────────────────────────────────────────────────────────
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Inicia sesión para continuar.'
login_manager.login_message_category = 'warning'


class User(UserMixin):
    def __init__(self, row):
        self.id     = str(row['id'])
        self.nombre = row['nombre']
        self.email  = row['email']
        self.rol    = row.get('rol', 'usuario')

    def is_admin(self):
        return self.rol == 'admin'


@login_manager.user_loader
def load_user(user_id):
    row = fetchone("SELECT * FROM usuarios WHERE id = %s AND activo = 1", (int(user_id),))
    return User(row) if row else None


@app.before_request
def require_login():
    public = {'login', 'setup', 'recuperar_contrasena', 'resetear_contrasena', 'static'}
    if request.endpoint not in public and not current_user.is_authenticated:
        return redirect(url_for('login', next=request.url))


ESTATUS_OPCIONES = ['Borrador', 'En progreso', 'En revisión', 'Completado', 'Cancelado']

# Parámetros de conexión PostgreSQL con keepalives para evitar SSL timeout de Neon
_PG_CONNECT_KWARGS = dict(
    keepalives=1,
    keepalives_idle=60,       # envía keepalive tras 60 s de inactividad
    keepalives_interval=10,   # reintenta cada 10 s
    keepalives_count=3,       # descarta la conexión tras 3 fallos
    connect_timeout=10,
) if use_pg else {}


# =============================================================================
#  Capa de conexión — abstrae PostgreSQL (Neon) vs SQLite (local)
# =============================================================================

def get_pg_conn():
    """Crea una conexión fresca a Neon con keepalives habilitados."""
    return psycopg2.connect(DATABASE_URL, **_PG_CONNECT_KWARGS)


def put_pg_conn(conn):
    """Cierra la conexión (sin pool — Neon gestiona pooling por su lado)."""
    try:
        conn.close()
    except Exception:
        pass


def get_sqlite_conn():
    """Conexión a SQLite local."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _pg_fetchall(sql, params):
    """SELECT con reconexión automática si la conexión SSL fue cerrada por Neon."""
    for intento in range(2):
        conn = get_pg_conn()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(sql, params)
                return [dict(r) for r in cur.fetchall()]
        except psycopg2.OperationalError:
            if intento == 1:
                raise          # falla definitiva en el 2.º intento
        finally:
            put_pg_conn(conn)


def _pg_fetchone(sql, params):
    for intento in range(2):
        conn = get_pg_conn()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(sql, params)
                row = cur.fetchone()
                return dict(row) if row else None
        except psycopg2.OperationalError:
            if intento == 1:
                raise
        finally:
            put_pg_conn(conn)


def _pg_execute(sql, params):
    for intento in range(2):
        conn = get_pg_conn()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params)
            conn.commit()
            return
        except psycopg2.OperationalError:
            try:
                conn.rollback()
            except Exception:
                pass
            if intento == 1:
                raise
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            raise
        finally:
            put_pg_conn(conn)


def fetchall(sql, params=()):
    """Ejecuta SELECT y devuelve lista de dicts."""
    if use_pg:
        return _pg_fetchall(sql, params)
    conn = get_sqlite_conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def fetchone(sql, params=()):
    """Ejecuta SELECT y devuelve un dict o None."""
    if use_pg:
        return _pg_fetchone(sql, params)
    conn = get_sqlite_conn()
    row = conn.execute(sql, params).fetchone()
    conn.close()
    return dict(row) if row else None


def execute(sql, params=()):
    """Ejecuta INSERT / UPDATE / DELETE."""
    if use_pg:
        return _pg_execute(sql, params)
    conn = get_sqlite_conn()
    conn.execute(sql, params)
    conn.commit()
    conn.close()


def ph(n=1):
    """
    Devuelve placeholders para la cantidad de parámetros indicada.
    PostgreSQL: %s, %s, ...   |   SQLite: ?, ?, ...
    """
    mark = '%s' if use_pg else '?'
    return ', '.join([mark] * n)


def p():
    """Placeholder individual."""
    return '%s' if use_pg else '?'


# =============================================================================
#  Init / migración de base de datos (solo SQLite local)
# =============================================================================

SQLITE_SCHEMA = '''
CREATE TABLE IF NOT EXISTS responsables (
    id                     INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre                 TEXT    NOT NULL,
    capacidad_horas_semana INTEGER NOT NULL CHECK (capacidad_horas_semana > 0)
);
CREATE TABLE IF NOT EXISTS divisiones (
    id     INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre TEXT    NOT NULL UNIQUE
);
CREATE TABLE IF NOT EXISTS key_users (
    id     INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre TEXT    NOT NULL
);
CREATE TABLE IF NOT EXISTS entregables (
    id                   INTEGER PRIMARY KEY AUTOINCREMENT,
    folio                TEXT    NOT NULL,
    nombre               TEXT    NOT NULL,
    descripcion          TEXT,
    fecha_inicio         TEXT,
    fecha_fin_desarrollo TEXT,
    fecha_entrega        TEXT,
    fecha_publicacion    TEXT,
    frd_aplica           INTEGER NOT NULL DEFAULT 0,
    frd_url              TEXT,
    frd_firmado          INTEGER NOT NULL DEFAULT 0,
    frd_aceptacion       INTEGER NOT NULL DEFAULT 0,
    responsable_id       INTEGER REFERENCES responsables(id),
    horas_desarrollo     INTEGER,
    porcentaje_avance    INTEGER NOT NULL DEFAULT 0,
    necesita_devops      INTEGER NOT NULL DEFAULT 0,
    division_id          INTEGER REFERENCES divisiones(id),
    key_user_id          INTEGER REFERENCES key_users(id),
    estatus              TEXT    NOT NULL DEFAULT 'Borrador',
    creado_en            TEXT
);
CREATE TABLE IF NOT EXISTS comentarios (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    entregable_id  INTEGER NOT NULL REFERENCES entregables(id) ON DELETE CASCADE,
    comentario     TEXT    NOT NULL,
    fecha          TEXT    NOT NULL
);
'''

SQLITE_MIGRATIONS = [
    ('folio',                'ALTER TABLE entregables ADD COLUMN folio TEXT'),
    ('fecha_fin_desarrollo', 'ALTER TABLE entregables ADD COLUMN fecha_fin_desarrollo TEXT'),
    ('fecha_publicacion',    'ALTER TABLE entregables ADD COLUMN fecha_publicacion TEXT'),
    ('frd_aplica',           'ALTER TABLE entregables ADD COLUMN frd_aplica INTEGER DEFAULT 0'),
    ('frd_url',              'ALTER TABLE entregables ADD COLUMN frd_url TEXT'),
    ('frd_firmado',          'ALTER TABLE entregables ADD COLUMN frd_firmado INTEGER DEFAULT 0'),
    ('frd_aceptacion',       'ALTER TABLE entregables ADD COLUMN frd_aceptacion INTEGER DEFAULT 0'),
    ('porcentaje_avance',    'ALTER TABLE entregables ADD COLUMN porcentaje_avance INTEGER DEFAULT 0'),
    ('necesita_devops',      'ALTER TABLE entregables ADD COLUMN necesita_devops INTEGER DEFAULT 0'),
    ('estatus',              "ALTER TABLE entregables ADD COLUMN estatus TEXT DEFAULT 'Borrador'"),
]

SEED_DATA = {
    'responsables': [('Maria Perez', 40), ('Juan Ramirez', 35), ('Ana Gomez', 30)],
    'divisiones':   [('TI',), ('Finanzas',), ('Operaciones',)],
    'key_users':    [('Karla Lopez',), ('Roberto Diaz',), ('Sofia Torres',)],
}


def init_db():
    """Inicializa SQLite local con schema, migraciones y seed."""
    conn = get_sqlite_conn()
    conn.executescript(SQLITE_SCHEMA)
    conn.commit()

    cols = [r[1] for r in conn.execute("PRAGMA table_info(entregables)").fetchall()]
    for col, sql in SQLITE_MIGRATIONS:
        if col not in cols:
            conn.execute(sql)
    conn.commit()

    if conn.execute('SELECT COUNT(*) FROM responsables').fetchone()[0] == 0:
        conn.executemany('INSERT INTO responsables (nombre, capacidad_horas_semana) VALUES (?, ?)',
                         SEED_DATA['responsables'])
    if conn.execute('SELECT COUNT(*) FROM divisiones').fetchone()[0] == 0:
        conn.executemany('INSERT INTO divisiones (nombre) VALUES (?)', SEED_DATA['divisiones'])
    if conn.execute('SELECT COUNT(*) FROM key_users').fetchone()[0] == 0:
        conn.executemany('INSERT INTO key_users (nombre) VALUES (?)', SEED_DATA['key_users'])
    conn.commit()
    conn.close()


if not use_pg:
    init_db()


# =============================================================================
#  Funciones de acceso a datos
# =============================================================================

def get_catalog_responsables():
    return fetchall('SELECT * FROM responsables ORDER BY id')


def get_catalog_divisiones():
    return fetchall('SELECT * FROM divisiones ORDER BY id')


def get_catalog_key_users():
    return fetchall('SELECT * FROM key_users ORDER BY id')


def get_catalog_proyectos():
    return fetchall('SELECT * FROM proyectos WHERE activo = 1 ORDER BY nombre')


def get_entregables(today=None, max_date=None):
    if today is not None and max_date is None:
        return fetchall(
            f'SELECT * FROM entregables WHERE fecha_entrega = {p()} ORDER BY fecha_entrega',
            (today,)
        )
    if today is not None and max_date is not None:
        return fetchall(
            f'SELECT * FROM entregables WHERE fecha_entrega > {p()} AND fecha_entrega <= {p()} ORDER BY fecha_entrega',
            (today, max_date)
        )
    return fetchall('SELECT * FROM entregables ORDER BY fecha_entrega')


def get_entregable_by_id(id):
    return fetchone(f'SELECT * FROM entregables WHERE id = {p()}', (id,))


def get_comments(entregable_id):
    return fetchall(
        f'SELECT * FROM comentarios WHERE entregable_id = {p()} ORDER BY fecha DESC',
        (entregable_id,)
    )


def insert_entregable(payload):
    sql = f'''
        INSERT INTO entregables
            (folio, nombre, descripcion, fecha_inicio, fecha_fin_desarrollo, fecha_entrega,
             fecha_publicacion, frd_aplica, frd_url, frd_firmado, frd_aceptacion,
             responsable_id, horas_desarrollo, porcentaje_avance, necesita_devops,
             division_id, key_user_id, proyecto_id, estatus, gherkin, creado_en)
        VALUES ({ph(21)})
    '''
    execute(sql, (
        payload.get('folio'), payload.get('nombre'), payload.get('descripcion'),
        payload.get('fecha_inicio'), payload.get('fecha_fin_desarrollo'),
        payload.get('fecha_entrega'), payload.get('fecha_publicacion'),
        payload.get('frd_aplica'), payload.get('frd_url'),
        payload.get('frd_firmado'), payload.get('frd_aceptacion'),
        payload.get('responsable_id'), payload.get('horas_desarrollo'),
        payload.get('porcentaje_avance'), payload.get('necesita_devops'),
        payload.get('division_id'), payload.get('key_user_id'), payload.get('proyecto_id'),
        payload.get('estatus', 'Borrador'), payload.get('gherkin'),
        payload.get('creado_en'),
    ))


def update_entregable(id, payload):
    mk = p()  # placeholder
    sql = f'''
        UPDATE entregables SET
            folio={mk}, nombre={mk}, descripcion={mk},
            fecha_inicio={mk}, fecha_fin_desarrollo={mk},
            fecha_entrega={mk}, fecha_publicacion={mk},
            frd_aplica={mk}, frd_url={mk}, frd_firmado={mk}, frd_aceptacion={mk},
            responsable_id={mk}, horas_desarrollo={mk}, porcentaje_avance={mk},
            necesita_devops={mk}, division_id={mk}, key_user_id={mk}, proyecto_id={mk},
            estatus={mk}, gherkin={mk}
        WHERE id={mk}
    '''
    execute(sql, (
        payload.get('folio'), payload.get('nombre'), payload.get('descripcion'),
        payload.get('fecha_inicio'), payload.get('fecha_fin_desarrollo'),
        payload.get('fecha_entrega'), payload.get('fecha_publicacion'),
        payload.get('frd_aplica'), payload.get('frd_url'),
        payload.get('frd_firmado'), payload.get('frd_aceptacion'),
        payload.get('responsable_id'), payload.get('horas_desarrollo'),
        payload.get('porcentaje_avance'), payload.get('necesita_devops'),
        payload.get('division_id'), payload.get('key_user_id'), payload.get('proyecto_id'),
        payload.get('estatus', 'Borrador'), payload.get('gherkin'), id,
    ))


def delete_entregable(id):
    execute('DELETE FROM entregables WHERE id = %s', (id,))


def insert_comment(entregable_id, comentario):
    execute(
        'INSERT INTO comentarios (entregable_id, comentario) VALUES (%s, %s)',
        (entregable_id, comentario)
    )


def insert_catalog(table, payload):
    if table == 'responsables':
        execute(
            f'INSERT INTO responsables (nombre, capacidad_horas_semana) VALUES ({ph(2)})',
            (payload['nombre'], payload['capacidad_horas_semana'])
        )
    elif table == 'divisiones':
        prefijo = payload.get('prefijo') or payload['nombre'][:3].upper()
        try:
            execute(f'INSERT INTO divisiones (nombre, prefijo) VALUES ({ph(2)})',
                    (payload['nombre'], prefijo))
        except Exception:
            # Fallback si la columna prefijo aún no existe en el schema
            execute(f'INSERT INTO divisiones (nombre) VALUES ({p()})', (payload['nombre'],))
    elif table == 'key_users':
        execute(f'INSERT INTO key_users (nombre) VALUES ({p()})', (payload['nombre'],))


# =============================================================================
#  Lógica de negocio
# =============================================================================

def parse_date(str_date):
    if str_date is None:
        return None
    if isinstance(str_date, datetime):
        return str_date.date()
    if isinstance(str_date, date):
        return str_date
    try:
        return datetime.strptime(str(str_date), '%Y-%m-%d').date()
    except Exception:
        return None


def calculate_estimated_entrega(fecha_inicio, horas_desarrollo, capacidad_sem):
    if not fecha_inicio or capacidad_sem <= 0:
        return None
    dias = int(round((horas_desarrollo / capacidad_sem) * 7))
    return fecha_inicio + timedelta(days=dias)


def get_delivery_status(entregable):
    hoy = date.today()
    fecha_entrega = parse_date(entregable.get('fecha_entrega'))
    if not fecha_entrega:
        return 'Sin fecha'
    avance = int(entregable.get('porcentaje_avance') or 0)
    if hoy > fecha_entrega:
        return 'Atrasado'
    if avance >= 80:
        return 'A tiempo'
    if (fecha_entrega - hoy).days <= 3:
        return 'En riesgo'
    return 'A tiempo'


def get_traffic_light(entregable):
    hoy = date.today()
    fecha_entrega = parse_date(entregable.get('fecha_entrega'))
    avance = int(entregable.get('porcentaje_avance') or 0)
    if fecha_entrega and hoy > fecha_entrega:
        return 'rojo'
    if avance < 40:
        return 'rojo'
    if avance < 80:
        return 'amarillo'
    return 'verde'


def status_for_entregable(entregable):
    fecha_entrega = parse_date(entregable.get('fecha_entrega'))
    fecha_inicio  = parse_date(entregable.get('fecha_inicio'))
    if not fecha_entrega or not fecha_inicio:
        return 'Sin fechas'
    resp = next((r for r in get_catalog_responsables() if r['id'] == entregable.get('responsable_id')), None)
    capacidad = resp['capacidad_horas_semana'] if resp else 0
    estimada = calculate_estimated_entrega(fecha_inicio, entregable.get('horas_desarrollo') or 0, capacidad)
    if not estimada:
        return 'Falta información'
    return 'En tiempo' if estimada <= fecha_entrega else 'Retardo probable'


def enrich_entregables(entregables):
    responsables = {r['id']: r['nombre'] for r in get_catalog_responsables()}
    divisiones   = {d['id']: d['nombre'] for d in get_catalog_divisiones()}
    key_users    = {k['id']: k['nombre'] for k in get_catalog_key_users()}
    proyectos    = {p['id']: p['nombre'] for p in get_catalog_proyectos()}
    enriched = []
    for e in entregables:
        r = dict(e)
        r['responsable']      = responsables.get(r.get('responsable_id'))
        r['division']         = divisiones.get(r.get('division_id'))
        r['key_user']         = key_users.get(r.get('key_user_id'))
        r['proyecto']         = proyectos.get(r.get('proyecto_id'))
        r['va_a_tiempo']      = get_delivery_status(r)
        r['semaforo_cliente'] = get_traffic_light(r)
        enriched.append(r)
    return enriched


def build_10_day_calendar(entregables):
    hoy = date.today()
    by_date = {}
    for e in entregables:
        fe = parse_date(e.get('fecha_entrega'))
        if fe:
            by_date.setdefault(fe.isoformat(), []).append(e)
    dias = []
    for i in range(10):
        dia = hoy + timedelta(days=i - 4)
        dias.append({
            'fecha': dia,
            'label': dia.strftime('%a %d %b'),
            'entregables': by_date.get(dia.isoformat(), []),
        })
    return dias


def count_by_division(entregables):
    counts = {}
    for e in entregables:
        div = e.get('division') or 'Sin división'
        counts[div] = counts.get(div, 0) + 1
    labels = list(counts.keys())
    return labels, [counts[k] for k in labels]


def compute_frd_metrics(entregables):
    total_frd = pendientes_frd = pendientes_firma = pendientes_aceptacion = completados_frd = 0
    for e in entregables:
        if not int(e.get('frd_aplica') or 0):
            continue
        total_frd += 1
        tiene_url = bool(e.get('frd_url'))
        firmado   = int(e.get('frd_firmado') or 0)
        aceptacion = int(e.get('frd_aceptacion') or 0)
        avance    = int(e.get('porcentaje_avance') or 0)
        if not tiene_url:
            pendientes_frd += 1
        elif not firmado:
            pendientes_firma += 1
        elif avance >= 100 and not aceptacion:
            pendientes_aceptacion += 1
        elif aceptacion:
            completados_frd += 1
    frd_pct = int((completados_frd / total_frd) * 100) if total_frd > 0 else 0
    return pendientes_frd, pendientes_firma, pendientes_aceptacion, frd_pct, total_frd


def count_semaforo(entregables):
    verde    = sum(1 for e in entregables if e.get('semaforo_cliente') == 'verde')
    amarillo = sum(1 for e in entregables if e.get('semaforo_cliente') == 'amarillo')
    rojo     = sum(1 for e in entregables if e.get('semaforo_cliente') == 'rojo')
    return verde, amarillo, rojo


# =============================================================================
#  Rutas Flask
# =============================================================================

def _form_payload(form):
    """Extrae y normaliza el payload del formulario (folio se genera automáticamente)."""
    return {
        'proyecto_id':         int(form.get('proyecto_id')) if form.get('proyecto_id') else None,
        'nombre':              form.get('nombre', '').strip(),
        'descripcion':         form.get('descripcion', '').strip(),
        'fecha_inicio':        form.get('fecha_inicio', '').strip() or None,
        'fecha_fin_desarrollo':form.get('fecha_fin_desarrollo', '').strip() or None,
        'fecha_entrega':       form.get('fecha_entrega', '').strip() or None,
        'fecha_publicacion':   form.get('fecha_publicacion', '').strip() or None,
        'frd_aplica':          1 if form.get('frd_aplica') == 'on' else 0,
        'frd_url':             form.get('frd_url', '').strip() or None,
        'frd_firmado':         1 if form.get('frd_firmado') == 'on' else 0,
        'frd_aceptacion':      1 if form.get('frd_aceptacion') == 'on' else 0,
        'responsable_id':      int(form.get('responsable_id')) if form.get('responsable_id') else None,
        'horas_desarrollo':    int(form.get('horas_desarrollo') or 0),
        'porcentaje_avance':   int(form.get('porcentaje_avance') or 0),
        'necesita_devops':     1 if form.get('necesita_devops') == 'on' else 0,
        'division_id':         int(form.get('division_id')) if form.get('division_id') else None,
        'key_user_id':         int(form.get('key_user_id')) if form.get('key_user_id') else None,
        'estatus':             form.get('estatus', 'Borrador'),
        'gherkin':             form.get('gherkin', '').strip() or None,
    }


def _catalogs():
    return (get_catalog_responsables(), get_catalog_divisiones(),
            get_catalog_key_users(), get_catalog_proyectos())


# =============================================================================
#  Folio automático
# =============================================================================

def _division_prefix(division_id):
    """Devuelve el prefijo de 3 letras de la división (columna prefijo o primeras 3 del nombre)."""
    if not division_id:
        return None
    div = fetchone(f"SELECT * FROM divisiones WHERE id = {p()}", (division_id,))
    if not div:
        return None
    pref = (div.get('prefijo') or '').strip().upper()
    return pref if pref else div['nombre'][:3].upper()


def generate_folio(division_id):
    """Genera el siguiente folio: PREFIX-NNN-AÑO basado en los existentes de esa división.
    Si no hay división asignada usa el prefijo GEN como genérico."""
    import re
    prefijo = _division_prefix(division_id) or 'GEN'
    year = date.today().year
    if division_id:
        rows = fetchall(
            f"SELECT folio FROM entregables WHERE division_id = {p()} AND folio IS NOT NULL",
            (division_id,)
        )
    else:
        # Sin división: buscar todos los folios GEN para calcular el siguiente
        rows = fetchall(
            f"SELECT folio FROM entregables WHERE folio LIKE {p()} AND division_id IS NULL",
            (f'{prefijo}-%',)
        )
    max_seq = 0
    for r in rows:
        m = re.match(rf'^{re.escape(prefijo)}-(\d+)', r.get('folio') or '', re.IGNORECASE)
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return f"{prefijo}-{max_seq + 1:03d}-{year}"


@app.route('/')
def dashboard():
    hoy             = date.today()
    hoy_str         = hoy.isoformat()
    limite_str      = (hoy + timedelta(days=7)).isoformat()
    division_filtro = request.args.get('division', '').strip()

    # Una sola consulta; todo el filtrado ocurre en Python
    all_entregables = enrich_entregables(get_entregables())
    divisiones      = get_catalog_divisiones()

    if division_filtro:
        filtered = [e for e in all_entregables if (e.get('division') or '') == division_filtro]
    else:
        filtered = all_entregables

    def _fecha_str(e):
        fd = e.get('fecha_entrega')
        if fd is None:
            return None
        return fd.isoformat()[:10] if isinstance(fd, (date, datetime)) else str(fd)[:10]

    pendientes_hoy = [e for e in filtered if _fecha_str(e) == hoy_str]
    proximos       = [e for e in filtered if _fecha_str(e) and hoy_str < _fecha_str(e) <= limite_str]
    calendario     = build_10_day_calendar(filtered)
    division_labels, division_values = count_by_division(filtered)
    pendientes_frd, pendientes_firma, pendientes_aceptacion, frd_pct, total_frd = compute_frd_metrics(filtered)
    sem_verde, sem_amarillo, sem_rojo = count_semaforo(filtered)

    vencidos_list = [
        e for e in filtered
        if parse_date(e.get('fecha_entrega')) and parse_date(e.get('fecha_entrega')) < hoy
        and (e.get('estatus') or 'Borrador') not in ('Completado', 'Cancelado')
    ]
    vencidos_list.sort(key=lambda e: parse_date(e.get('fecha_entrega')))

    return render_template(
        'dashboard.html',
        now=datetime.now(),
        pendientes_hoy=pendientes_hoy,
        proximos=proximos,
        calendario=calendario,
        division_labels=division_labels,
        division_values=division_values,
        pendientes_frd=pendientes_frd,
        pendientes_firma=pendientes_firma,
        pendientes_aceptacion=pendientes_aceptacion,
        frd_pct=frd_pct,
        total_frd=total_frd,
        sem_verde=sem_verde,
        sem_amarillo=sem_amarillo,
        sem_rojo=sem_rojo,
        total_entregables=len(filtered),
        divisiones=divisiones,
        division_seleccionada=division_filtro,
        vencidos_list=vencidos_list,
        vencidos_count=len(vencidos_list),
    )


@app.route('/entregables')
def entregables():
    items        = enrich_entregables(get_entregables())
    divisiones   = get_catalog_divisiones()
    responsables = get_catalog_responsables()
    proyectos    = get_catalog_proyectos()
    return render_template('entregables.html', entregables=items,
                           divisiones=divisiones, responsables=responsables,
                           proyectos=proyectos)


@app.route('/entregables/nuevo', methods=['GET', 'POST'])
def nuevo_entregable():
    responsables, divisiones, key_users, proyectos = _catalogs()
    if request.method == 'POST':
        payload = _form_payload(request.form)
        if not payload['nombre'] or not payload['responsable_id']:
            flash('Nombre y responsable son obligatorios.', 'error')
        else:
            payload['folio']     = generate_folio(payload.get('division_id'))
            payload['creado_en'] = datetime.now().isoformat()
            insert_entregable(payload)
            flash(f'Entregable creado con folio {payload["folio"] or "sin folio"}.', 'success')
            return redirect(url_for('entregables'))
    return render_template('create_entregable.html', responsables=responsables,
                           divisiones=divisiones, key_users=key_users, proyectos=proyectos,
                           estatus_opciones=ESTATUS_OPCIONES, entregable=None)


@app.route('/entregables/<int:id>/editar', methods=['GET', 'POST'])
def editar_entregable(id):
    e = get_entregable_by_id(id)
    if not e:
        return 'Entregable no encontrado', 404
    responsables, divisiones, key_users, proyectos = _catalogs()
    if request.method == 'POST':
        payload = _form_payload(request.form)
        payload['folio'] = e['folio']  # folio inmutable: conservar el original
        if not payload['nombre'] or not payload['responsable_id']:
            flash('Nombre y responsable son obligatorios.', 'error')
        else:
            update_entregable(id, payload)
            flash('Entregable actualizado correctamente.', 'success')
            return redirect(url_for('detalle_entregable', id=id))
    return render_template('create_entregable.html', responsables=responsables,
                           divisiones=divisiones, key_users=key_users, proyectos=proyectos,
                           estatus_opciones=ESTATUS_OPCIONES, entregable=e)


@app.route('/entregables/<int:id>')
def detalle_entregable(id):
    e = get_entregable_by_id(id)
    if not e:
        return 'Entregable no encontrado', 404
    responsable = next((r for r in get_catalog_responsables() if r['id'] == e.get('responsable_id')), None)
    e['responsable']           = responsable['nombre'] if responsable else None
    e['capacidad_horas_semana']= responsable['capacidad_horas_semana'] if responsable else None
    e['division']  = next((d['nombre'] for d in get_catalog_divisiones()  if d['id'] == e.get('division_id')), None)
    e['key_user']  = next((k['nombre'] for k in get_catalog_key_users()   if k['id'] == e.get('key_user_id')), None)
    e['va_a_tiempo']      = get_delivery_status(e)
    e['semaforo_cliente'] = get_traffic_light(e)
    comentarios = get_comments(id)
    estado   = status_for_entregable(e)
    sugerida = None
    if e.get('fecha_inicio') and e.get('horas_desarrollo') and e.get('capacidad_horas_semana'):
        sugerida = calculate_estimated_entrega(
            parse_date(e['fecha_inicio']), e['horas_desarrollo'], e['capacidad_horas_semana']
        )
    # Registros de trabajo (seguimiento de horas)
    registros_trabajo = fetchall(
        "SELECT * FROM registros_trabajo WHERE entregable_id = %s ORDER BY fecha DESC, id DESC",
        (id,)
    )
    horas_estimadas   = float(e.get('horas_desarrollo') or 0)
    horas_registradas = sum(float(r.get('horas_trabajadas') or 0) for r in registros_trabajo)
    horas_restantes   = max(0, horas_estimadas - horas_registradas)
    pct_horas         = min(100, round((horas_registradas / horas_estimadas * 100) if horas_estimadas > 0 else 0))

    # Contexto de navegación desde la cola
    cola_resp_id = request.args.get('cola_resp', type=int)
    cola_prev_id = cola_next_id = cola_position = cola_total = cola_resp_nombre = None
    if cola_resp_id:
        cola_ids = [
            r['id'] for r in fetchall("""
                SELECT id FROM entregables
                WHERE responsable_id = %s
                  AND COALESCE(estatus, 'Borrador') NOT IN ('Completado', 'Cancelado')
                ORDER BY orden_responsable ASC NULLS LAST, id ASC
            """, (cola_resp_id,))
        ]
        if id in cola_ids:
            pos          = cola_ids.index(id)
            cola_position = pos + 1
            cola_total    = len(cola_ids)
            cola_prev_id  = cola_ids[pos - 1] if pos > 0 else None
            cola_next_id  = cola_ids[pos + 1] if pos < len(cola_ids) - 1 else None
        resp_row = fetchone("SELECT nombre FROM responsables WHERE id = %s", (cola_resp_id,))
        cola_resp_nombre = resp_row['nombre'] if resp_row else None

    return render_template('detalle_entregable.html', e=e, comentarios=comentarios,
                           estado=estado, sugerida=sugerida,
                           now=datetime.now(),
                           registros_trabajo=registros_trabajo,
                           horas_estimadas=horas_estimadas,
                           horas_registradas=horas_registradas,
                           horas_restantes=horas_restantes,
                           pct_horas=pct_horas,
                           cola_resp_id=cola_resp_id,
                           cola_prev_id=cola_prev_id,
                           cola_next_id=cola_next_id,
                           cola_position=cola_position,
                           cola_total=cola_total,
                           cola_resp_nombre=cola_resp_nombre)


@app.route('/entregables/<int:id>/comentarios', methods=['POST'])
def agregar_comentario(id):
    texto = request.form.get('comentario', '').strip()
    if not texto:
        flash('El comentario no puede estar vacío.', 'error')
        return redirect(url_for('detalle_entregable', id=id))
    insert_comment(id, texto)
    flash('Comentario agregado.', 'success')
    return redirect(url_for('detalle_entregable', id=id))


@app.route('/entregables/<int:id>/eliminar', methods=['POST'])
def eliminar_entregable(id):
    e = get_entregable_by_id(id)
    if not e:
        return 'Entregable no encontrado', 404
    delete_entregable(id)
    flash(f'Entregable "{e["nombre"]}" eliminado.', 'success')
    return redirect(url_for('entregables'))


@app.route('/dashboard/export/<fmt>')
def dashboard_export(fmt):
    if fmt not in ('excel', 'pdf', 'pptx'):
        return 'Formato no soportado', 400

    from exports import generate_excel, generate_pdf, generate_pptx

    division_filtro = request.args.get('division', '').strip()
    all_entregables = enrich_entregables(get_entregables())
    filtered = [e for e in all_entregables if (e.get('division') or '') == division_filtro] \
               if division_filtro else all_entregables

    hoy = date.today()
    vencidos_list = [
        e for e in filtered
        if parse_date(e.get('fecha_entrega')) and parse_date(e.get('fecha_entrega')) < hoy
        and (e.get('estatus') or '') not in ('Completado', 'Cancelado')
    ]

    pendientes_frd, pendientes_firma, pendientes_aceptacion, frd_pct, total_frd = compute_frd_metrics(filtered)
    sem_verde, sem_amarillo, sem_rojo = count_semaforo(filtered)

    kwargs = dict(
        filtered=filtered,
        division_filtro=division_filtro,
        sem_verde=sem_verde, sem_amarillo=sem_amarillo, sem_rojo=sem_rojo,
        pendientes_frd=pendientes_frd, pendientes_firma=pendientes_firma,
        pendientes_aceptacion=pendientes_aceptacion,
        frd_pct=frd_pct, total_frd=total_frd,
    )

    div_slug = division_filtro.replace(' ', '_') if division_filtro else 'todas'
    fecha_slug = datetime.now().strftime('%Y%m%d')

    if fmt == 'excel':
        buf = generate_excel(**kwargs)
        return send_file(buf, download_name=f'entregables_{div_slug}_{fecha_slug}.xlsx',
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        buf = generate_pdf(**kwargs)
        return send_file(buf, download_name=f'entregables_{div_slug}_{fecha_slug}.pdf',
                         as_attachment=True, mimetype='application/pdf')
    else:
        buf = generate_pptx(**kwargs)
        return send_file(buf, download_name=f'entregables_{div_slug}_{fecha_slug}.pptx',
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation')


@app.route('/catalogos/responsables', methods=['GET', 'POST'])
def catalogo_responsables():
    if request.method == 'POST':
        nombre   = request.form.get('nombre', '').strip()
        capacidad = int(request.form.get('capacidad_horas_semana') or 0)
        if nombre and capacidad > 0:
            insert_catalog('responsables', {'nombre': nombre, 'capacidad_horas_semana': capacidad})
            flash('Responsable agregado.', 'success')
        else:
            flash('Nombre y capacidad válidos son obligatorios.', 'error')
    return render_template('responsables.html', responsables=get_catalog_responsables())


@app.route('/api/folio-preview')
def api_folio_preview():
    """Devuelve el siguiente folio para una división (AJAX)."""
    division_id = request.args.get('division_id', type=int)
    folio = generate_folio(division_id) if division_id else None
    return jsonify({'folio': folio or '—'})


@app.route('/catalogos/divisiones', methods=['GET', 'POST'])
def catalogo_divisiones():
    if request.method == 'POST':
        nombre  = request.form.get('nombre', '').strip()
        prefijo = request.form.get('prefijo', '').strip().upper()[:5]
        if not prefijo:
            prefijo = nombre[:3].upper()
        if nombre:
            insert_catalog('divisiones', {'nombre': nombre, 'prefijo': prefijo})
            flash('División agregada.', 'success')
        else:
            flash('Nombre es obligatorio.', 'error')
    return render_template('divisiones.html', divisiones=get_catalog_divisiones())


@app.route('/catalogos/key_users', methods=['GET', 'POST'])
def catalogo_key_users():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        if nombre:
            insert_catalog('key_users', {'nombre': nombre})
            flash('Key user agregado.', 'success')
        else:
            flash('Nombre es obligatorio.', 'error')
    return render_template('key_users.html', key_users=get_catalog_key_users())


@app.route('/catalogos/proyectos', methods=['GET', 'POST'])
def catalogo_proyectos():
    if request.method == 'POST':
        nombre      = request.form.get('nombre', '').strip()
        descripcion = request.form.get('descripcion', '').strip() or None
        division_id = request.form.get('division_id') or None
        if nombre:
            execute(
                f"INSERT INTO proyectos (nombre, descripcion, division_id) VALUES ({ph(3)})",
                (nombre, descripcion, division_id)
            )
            flash(f'Proyecto "{nombre}" creado.', 'success')
        else:
            flash('El nombre del proyecto es obligatorio.', 'error')
    proyectos  = fetchall("""
        SELECT p.*, d.nombre AS division_nombre,
               COUNT(e.id) AS total_entregables
        FROM proyectos p
        LEFT JOIN divisiones d ON p.division_id = d.id
        LEFT JOIN entregables e ON e.proyecto_id = p.id
        WHERE p.activo = 1
        GROUP BY p.id, d.nombre
        ORDER BY p.nombre
    """)
    return render_template('proyectos.html',
                           proyectos=proyectos,
                           divisiones=get_catalog_divisiones())


@app.route('/catalogos/proyectos/<int:id>/eliminar', methods=['POST'])
def eliminar_proyecto(id):
    execute(f"UPDATE proyectos SET activo = 0 WHERE id = {p()}", (id,))
    flash('Proyecto eliminado.', 'success')
    return redirect(url_for('catalogo_proyectos'))


# =============================================================================
#  Seguimiento de horas
# =============================================================================

@app.route('/entregables/<int:id>/horas', methods=['POST'])
def agregar_horas(id):
    fecha  = request.form.get('fecha', '').strip() or date.today().isoformat()
    horas  = request.form.get('horas_trabajadas', '').strip()
    notas  = request.form.get('notas', '').strip() or None
    try:
        horas_f = float(horas)
        if horas_f <= 0:
            raise ValueError
    except (ValueError, TypeError):
        flash('Ingresa horas válidas (número mayor a 0).', 'error')
        return redirect(url_for('detalle_entregable', id=id))
    execute(
        "INSERT INTO registros_trabajo (entregable_id, fecha, horas_trabajadas, notas) "
        "VALUES (%s, %s, %s, %s)",
        (id, fecha, horas_f, notas)
    )
    flash('Horas registradas.', 'success')
    return redirect(url_for('detalle_entregable', id=id))


@app.route('/entregables/<int:id>/horas/<int:reg_id>/eliminar', methods=['POST'])
def eliminar_registro_horas(id, reg_id):
    execute(
        "DELETE FROM registros_trabajo WHERE id = %s AND entregable_id = %s",
        (reg_id, id)
    )
    flash('Registro eliminado.', 'success')
    return redirect(url_for('detalle_entregable', id=id))


# =============================================================================
#  Cola por responsable
# =============================================================================

@app.route('/cola')
def cola_responsables():
    proyecto_id = request.args.get('proyecto_id', type=int)
    responsables = get_catalog_responsables()
    proyectos    = get_catalog_proyectos()
    cola = []
    for r in responsables:
        filtro_proyecto = f"AND e.proyecto_id = {p()}" if proyecto_id else ""
        params = (r['id'], proyecto_id) if proyecto_id else (r['id'],)
        ents = enrich_entregables(fetchall(f"""
            SELECT e.* FROM entregables e
            LEFT JOIN proyectos pr ON e.proyecto_id = pr.id
            WHERE e.responsable_id = {p()}
              AND COALESCE(e.estatus, 'Borrador') NOT IN ('Completado', 'Cancelado')
              {filtro_proyecto}
            ORDER BY COALESCE(pr.nombre, 'zzz') ASC,
                     e.orden_responsable ASC NULLS LAST,
                     e.id ASC
        """, params))
        if ents:
            cola.append({'responsable': r, 'entregables': ents})
    return render_template('cola.html', cola=cola,
                           proyectos=proyectos,
                           proyecto_sel=proyecto_id)


@app.route('/cola/<int:responsable_id>')
def cola_responsable(responsable_id):
    """Redirige al primer entregable activo de la cola del responsable."""
    row = fetchone("""
        SELECT id FROM entregables
        WHERE responsable_id = %s
          AND COALESCE(estatus, 'Borrador') NOT IN ('Completado', 'Cancelado')
        ORDER BY orden_responsable ASC NULLS LAST, id ASC
        LIMIT 1
    """, (responsable_id,))
    if not row:
        flash('Este responsable no tiene tareas activas en la cola.', 'warning')
        return redirect(url_for('cola_responsables'))
    return redirect(url_for('detalle_entregable', id=row['id'], cola_resp=responsable_id))


@app.route('/cola/reordenar', methods=['POST'])
def reordenar_cola():
    data           = request.get_json(force=True)
    responsable_id = data.get('responsable_id')
    orden          = data.get('orden', [])
    for i, ent_id in enumerate(orden, 1):
        execute(
            "UPDATE entregables SET orden_responsable = %s "
            "WHERE id = %s AND responsable_id = %s",
            (i, ent_id, responsable_id)
        )
    return {'ok': True}


# =============================================================================
#  OKR — Funciones de base de datos
# =============================================================================

def get_ciclos_okr():
    return fetchall("SELECT * FROM ciclos_okr ORDER BY fecha_inicio DESC")


def get_ciclo_by_id(ciclo_id):
    return fetchone("SELECT * FROM ciclos_okr WHERE id = %s", (ciclo_id,))


def get_objetivos_by_ciclo(ciclo_id):
    return fetchall("""
        SELECT o.*, r.nombre AS responsable_nombre
        FROM objetivos o
        LEFT JOIN responsables r ON o.responsable_id = r.id
        WHERE o.ciclo_id = %s
        ORDER BY o.id
    """, (ciclo_id,))


def get_objetivo_by_id(obj_id):
    return fetchone("""
        SELECT o.*, r.nombre AS responsable_nombre,
               c.nombre AS ciclo_nombre, c.id AS ciclo_id_ref
        FROM objetivos o
        LEFT JOIN responsables r ON o.responsable_id = r.id
        LEFT JOIN ciclos_okr c   ON o.ciclo_id = c.id
        WHERE o.id = %s
    """, (obj_id,))


def get_krs_by_objetivo(obj_id):
    return fetchall("""
        SELECT * FROM key_results
        WHERE objetivo_id = %s
        ORDER BY id
    """, (obj_id,))


def get_entregables_by_kr(kr_id):
    return fetchall("""
        SELECT e.*, ekr.contribucion_pct, r.nombre AS responsable
        FROM entregable_key_result ekr
        JOIN entregables e ON ekr.entregable_id = e.id
        LEFT JOIN responsables r ON e.responsable_id = r.id
        WHERE ekr.key_result_id = %s
        ORDER BY e.id
    """, (kr_id,))


def calc_kr_progreso(kr, entregables):
    """Calcula 0-100 para un KR basado en sus entregables vinculados."""
    if entregables:
        total_peso = sum(e.get('contribucion_pct') or 100 for e in entregables)
        if total_peso > 0:
            ponderado = sum(
                (e.get('porcentaje_avance') or 0) * (e.get('contribucion_pct') or 100)
                for e in entregables
            ) / total_peso
            return round(ponderado)
    # Fallback: valor manual
    rango = float(kr.get('valor_meta') or 100) - float(kr.get('valor_inicial') or 0)
    if rango > 0:
        actual  = float(kr.get('valor_actual')  or 0)
        inicial = float(kr.get('valor_inicial') or 0)
        return min(100, round(((actual - inicial) / rango) * 100))
    return 0


def enrich_objetivo(obj):
    krs = get_krs_by_objetivo(obj['id'])
    for kr in krs:
        kr['entregables'] = enrich_entregables(get_entregables_by_kr(kr['id']))
        kr['progreso'] = calc_kr_progreso(kr, kr['entregables'])
        kr['semaforo'] = 'verde' if kr['progreso'] >= 70 else ('amarillo' if kr['progreso'] >= 40 else 'rojo')
    obj['key_results'] = krs
    obj['progreso'] = round(sum(k['progreso'] for k in krs) / len(krs)) if krs else 0
    obj['semaforo'] = 'verde' if obj['progreso'] >= 70 else ('amarillo' if obj['progreso'] >= 40 else 'rojo')
    return obj


# =============================================================================
#  OKR — Rutas
# =============================================================================

OKR_COLORS = ['#4F46E5', '#0284C7', '#16A34A', '#D97706',
              '#DC2626', '#7C3AED', '#DB2777', '#0891B2']


@app.route('/okr')
def okr():
    ciclos = get_ciclos_okr()
    ciclo_id = request.args.get('ciclo', type=int)
    if ciclo_id:
        ciclo_activo = get_ciclo_by_id(ciclo_id)
    else:
        ciclo_activo = fetchone(
            "SELECT * FROM ciclos_okr WHERE activo = 1 ORDER BY fecha_inicio DESC LIMIT 1"
        )
        if not ciclo_activo and ciclos:
            ciclo_activo = ciclos[0]

    objetivos = []
    if ciclo_activo:
        for obj in get_objetivos_by_ciclo(ciclo_activo['id']):
            objetivos.append(enrich_objetivo(obj))

    responsables = get_catalog_responsables()
    return render_template('okr.html',
        ciclos=ciclos,
        ciclo_activo=ciclo_activo,
        objetivos=objetivos,
        responsables=responsables,
        okr_colors=OKR_COLORS,
    )


@app.route('/okr/ciclos', methods=['POST'])
def crear_ciclo():
    nombre       = request.form.get('nombre', '').strip()
    fecha_inicio = request.form.get('fecha_inicio', '').strip()
    fecha_fin    = request.form.get('fecha_fin', '').strip()
    hacer_activo = request.form.get('activo') == 'on'
    if not nombre or not fecha_inicio or not fecha_fin:
        flash('Nombre, fecha inicio y fecha fin son obligatorios.', 'error')
        return redirect(url_for('okr'))
    if hacer_activo:
        execute("UPDATE ciclos_okr SET activo = 0")
    execute(
        "INSERT INTO ciclos_okr (nombre, fecha_inicio, fecha_fin, activo) VALUES (%s,%s,%s,%s)",
        (nombre, fecha_inicio, fecha_fin, 1 if hacer_activo else 0)
    )
    flash(f'Ciclo "{nombre}" creado.', 'success')
    return redirect(url_for('okr'))


@app.route('/okr/ciclos/<int:id>/activar', methods=['POST'])
def activar_ciclo(id):
    execute("UPDATE ciclos_okr SET activo = 0")
    execute("UPDATE ciclos_okr SET activo = 1 WHERE id = %s", (id,))
    flash('Ciclo activado.', 'success')
    return redirect(url_for('okr'))


@app.route('/okr/ciclos/<int:id>/eliminar', methods=['POST'])
def eliminar_ciclo(id):
    execute("DELETE FROM ciclos_okr WHERE id = %s", (id,))
    flash('Ciclo eliminado.', 'success')
    return redirect(url_for('okr'))


@app.route('/okr/objetivos', methods=['POST'])
def crear_objetivo():
    ciclo_id       = request.form.get('ciclo_id', type=int)
    titulo         = request.form.get('titulo', '').strip()
    descripcion    = request.form.get('descripcion', '').strip() or None
    responsable_id = request.form.get('responsable_id') or None
    color          = request.form.get('color', '#4F46E5')
    if not ciclo_id or not titulo:
        flash('Ciclo y título son obligatorios.', 'error')
        return redirect(url_for('okr'))
    execute(
        "INSERT INTO objetivos (ciclo_id, titulo, descripcion, responsable_id, color) "
        "VALUES (%s,%s,%s,%s,%s)",
        (ciclo_id, titulo, descripcion, responsable_id, color)
    )
    flash(f'Objetivo "{titulo}" creado.', 'success')
    return redirect(url_for('okr', ciclo=ciclo_id))


@app.route('/okr/objetivos/<int:id>')
def detalle_objetivo(id):
    obj = get_objetivo_by_id(id)
    if not obj:
        return 'Objetivo no encontrado', 404
    obj = enrich_objetivo(obj)
    all_entregables = enrich_entregables(get_entregables())
    return render_template('okr_objetivo.html',
        obj=obj,
        all_entregables=all_entregables,
        responsables=get_catalog_responsables(),
    )


@app.route('/okr/objetivos/<int:id>/eliminar', methods=['POST'])
def eliminar_objetivo(id):
    obj = fetchone("SELECT ciclo_id FROM objetivos WHERE id = %s", (id,))
    ciclo_id = obj['ciclo_id'] if obj else None
    execute("DELETE FROM objetivos WHERE id = %s", (id,))
    flash('Objetivo eliminado.', 'success')
    return redirect(url_for('okr', ciclo=ciclo_id) if ciclo_id else url_for('okr'))


@app.route('/okr/key-results', methods=['POST'])
def crear_kr():
    objetivo_id   = request.form.get('objetivo_id', type=int)
    descripcion   = request.form.get('descripcion', '').strip()
    metrica       = request.form.get('metrica', '').strip() or None
    valor_inicial = float(request.form.get('valor_inicial') or 0)
    valor_meta    = float(request.form.get('valor_meta') or 100)
    unidad        = request.form.get('unidad', '%').strip() or '%'
    if not objetivo_id or not descripcion:
        flash('Objetivo y descripción son obligatorios.', 'error')
        return redirect(url_for('okr'))
    execute(
        "INSERT INTO key_results (objetivo_id, descripcion, metrica, valor_inicial, valor_meta, unidad) "
        "VALUES (%s,%s,%s,%s,%s,%s)",
        (objetivo_id, descripcion, metrica, valor_inicial, valor_meta, unidad)
    )
    flash('Key Result creado.', 'success')
    return redirect(url_for('detalle_objetivo', id=objetivo_id))


@app.route('/okr/key-results/<int:id>/eliminar', methods=['POST'])
def eliminar_kr(id):
    kr = fetchone("SELECT objetivo_id FROM key_results WHERE id = %s", (id,))
    obj_id = kr['objetivo_id'] if kr else None
    execute("DELETE FROM key_results WHERE id = %s", (id,))
    flash('Key Result eliminado.', 'success')
    return redirect(url_for('detalle_objetivo', id=obj_id) if obj_id else url_for('okr'))


@app.route('/okr/key-results/<int:id>/actualizar', methods=['POST'])
def actualizar_kr(id):
    valor_actual = float(request.form.get('valor_actual') or 0)
    kr = fetchone("SELECT objetivo_id FROM key_results WHERE id = %s", (id,))
    execute("UPDATE key_results SET valor_actual = %s WHERE id = %s", (valor_actual, id))
    flash('Key Result actualizado.', 'success')
    return redirect(url_for('detalle_objetivo', id=kr['objetivo_id']) if kr else url_for('okr'))


@app.route('/okr/key-results/<int:kr_id>/vincular', methods=['POST'])
def vincular_entregable_kr(kr_id):
    entregable_id    = request.form.get('entregable_id', type=int)
    contribucion_pct = int(request.form.get('contribucion_pct') or 100)
    kr = fetchone("SELECT objetivo_id FROM key_results WHERE id = %s", (kr_id,))
    if not entregable_id or not kr:
        flash('Datos inválidos.', 'error')
        return redirect(url_for('okr'))
    execute(
        "INSERT INTO entregable_key_result (entregable_id, key_result_id, contribucion_pct) "
        "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
        (entregable_id, kr_id, contribucion_pct)
    )
    flash('Entregable vinculado al Key Result.', 'success')
    return redirect(url_for('detalle_objetivo', id=kr['objetivo_id']))


@app.route('/okr/key-results/<int:kr_id>/desvincular/<int:ent_id>', methods=['POST'])
def desvincular_entregable_kr(kr_id, ent_id):
    kr = fetchone("SELECT objetivo_id FROM key_results WHERE id = %s", (kr_id,))
    execute(
        "DELETE FROM entregable_key_result WHERE key_result_id=%s AND entregable_id=%s",
        (kr_id, ent_id)
    )
    flash('Entregable desvinculado.', 'success')
    return redirect(url_for('detalle_objetivo', id=kr['objetivo_id']) if kr else url_for('okr'))


# =============================================================================
#  Autenticación
# =============================================================================

def _send_reset_email(to_email, nombre, reset_url):
    """Envía correo de recuperación. Retorna True si se envió."""
    mail_server = os.environ.get('MAIL_SERVER', '')
    if not mail_server:
        return False
    try:
        port     = int(os.environ.get('MAIL_PORT', 587))
        username = os.environ.get('MAIL_USERNAME', '')
        password = os.environ.get('MAIL_PASSWORD', '')
        from_addr = os.environ.get('MAIL_FROM', username)

        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'Recuperación de contraseña — Gestor de Entregables'
        msg['From']    = from_addr
        msg['To']      = to_email

        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:480px;margin:auto">
          <h2 style="color:#4F46E5">Restablecer contraseña</h2>
          <p>Hola <strong>{nombre}</strong>,</p>
          <p>Haz clic en el botón para crear una nueva contraseña.
             El enlace expira en <strong>1 hora</strong>.</p>
          <p style="text-align:center;margin:32px 0">
            <a href="{reset_url}"
               style="background:#4F46E5;color:#fff;padding:12px 28px;
                      border-radius:8px;text-decoration:none;font-weight:bold">
              Restablecer contraseña
            </a>
          </p>
          <p style="color:#6B7280;font-size:.85rem">
            Si no solicitaste esto, ignora este mensaje.
          </p>
        </div>"""
        msg.attach(MIMEText(html, 'html'))

        with smtplib.SMTP(mail_server, port) as server:
            server.starttls()
            if username:
                server.login(username, password)
            server.sendmail(from_addr, to_email, msg.as_string())
        return True
    except Exception as ex:
        print(f'[MAIL ERROR] {ex}')
        return False


@app.route('/login', methods=['GET', 'POST'])
def login():
    total = fetchone("SELECT COUNT(*) AS n FROM usuarios")
    if total and total['n'] == 0:
        return redirect(url_for('setup'))

    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'

        row = fetchone(
            "SELECT * FROM usuarios WHERE LOWER(email) = %s AND activo = 1", (email,)
        )
        if row and check_password_hash(row['password_hash'], password):
            login_user(User(row), remember=remember)
            execute("UPDATE usuarios SET ultimo_acceso = NOW() WHERE id = %s", (row['id'],))
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Correo o contraseña incorrectos.', 'error')

    return render_template('login.html')


@app.route('/setup', methods=['GET', 'POST'])
def setup():
    """Solo disponible cuando no existen usuarios."""
    total = fetchone("SELECT COUNT(*) AS n FROM usuarios")
    if total and total['n'] > 0:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nombre   = request.form.get('nombre', '').strip()
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm', '')

        if not nombre or not email or not password:
            flash('Todos los campos son obligatorios.', 'error')
        elif password != confirm:
            flash('Las contraseñas no coinciden.', 'error')
        elif len(password) < 8:
            flash('La contraseña debe tener mínimo 8 caracteres.', 'error')
        else:
            execute(
                "INSERT INTO usuarios (nombre, email, password_hash, rol) VALUES (%s,%s,%s,'admin')",
                (nombre, email, generate_password_hash(password))
            )
            flash('Administrador creado. Inicia sesión.', 'success')
            return redirect(url_for('login'))

    return render_template('login.html', setup_mode=True)


@app.route('/logout')
def logout():
    logout_user()
    flash('Sesión cerrada correctamente.', 'success')
    return redirect(url_for('login'))


@app.route('/recuperar-contrasena', methods=['GET', 'POST'])
def recuperar_contrasena():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        row   = fetchone(
            "SELECT * FROM usuarios WHERE LOWER(email) = %s AND activo = 1", (email,)
        )
        if row:
            s         = URLSafeTimedSerializer(app.secret_key)
            token     = s.dumps(email, salt='reset-password')
            reset_url = url_for('resetear_contrasena', token=token, _external=True)
            sent      = _send_reset_email(email, row['nombre'], reset_url)
            if sent:
                flash('Enlace de recuperación enviado. Revisa tu correo.', 'success')
            else:
                # Fallback sin SMTP: el admin puede copiar el link
                flash(f'Correo no configurado. Enlace de recuperación: {reset_url}', 'warning')
        else:
            flash('Si ese correo está registrado recibirás un enlace de recuperación.', 'info')
        return redirect(url_for('login'))

    return render_template('recuperar_contrasena.html')


@app.route('/resetear-contrasena/<token>', methods=['GET', 'POST'])
def resetear_contrasena(token):
    s = URLSafeTimedSerializer(app.secret_key)
    try:
        email = s.loads(token, salt='reset-password', max_age=3600)
    except SignatureExpired:
        flash('El enlace expiró. Solicita uno nuevo.', 'error')
        return redirect(url_for('recuperar_contrasena'))
    except BadSignature:
        flash('Enlace inválido.', 'error')
        return redirect(url_for('recuperar_contrasena'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm', '')
        if len(password) < 8:
            flash('La contraseña debe tener mínimo 8 caracteres.', 'error')
        elif password != confirm:
            flash('Las contraseñas no coinciden.', 'error')
        else:
            execute(
                "UPDATE usuarios SET password_hash = %s WHERE LOWER(email) = %s",
                (generate_password_hash(password), email.lower())
            )
            flash('Contraseña actualizada. Inicia sesión.', 'success')
            return redirect(url_for('login'))

    return render_template('resetear_contrasena.html', token=token)


# =============================================================================
#  Gestión de usuarios
# =============================================================================

@app.route('/usuarios')
def gestion_usuarios():
    if not current_user.is_admin():
        flash('Solo los administradores pueden gestionar usuarios.', 'error')
        return redirect(url_for('dashboard'))
    users = fetchall("SELECT * FROM usuarios ORDER BY creado_en DESC")
    return render_template('usuarios.html', usuarios=users)


@app.route('/usuarios/nuevo', methods=['POST'])
def crear_usuario():
    if not current_user.is_admin():
        return redirect(url_for('dashboard'))
    nombre   = request.form.get('nombre', '').strip()
    email    = request.form.get('email', '').strip().lower()
    rol      = request.form.get('rol', 'usuario')
    password = request.form.get('password', '')

    if not nombre or not email or not password:
        flash('Nombre, correo y contraseña son obligatorios.', 'error')
    elif fetchone("SELECT id FROM usuarios WHERE LOWER(email) = %s", (email,)):
        flash('Ya existe un usuario con ese correo.', 'error')
    elif len(password) < 8:
        flash('La contraseña debe tener mínimo 8 caracteres.', 'error')
    else:
        execute(
            "INSERT INTO usuarios (nombre, email, password_hash, rol) VALUES (%s,%s,%s,%s)",
            (nombre, email, generate_password_hash(password), rol)
        )
        flash(f'Usuario {nombre} creado.', 'success')
    return redirect(url_for('gestion_usuarios'))


@app.route('/usuarios/<int:id>/toggle', methods=['POST'])
def toggle_usuario(id):
    if not current_user.is_admin():
        return redirect(url_for('dashboard'))
    if str(id) == current_user.id:
        flash('No puedes desactivar tu propio usuario.', 'error')
        return redirect(url_for('gestion_usuarios'))
    row = fetchone("SELECT activo FROM usuarios WHERE id = %s", (id,))
    if row:
        execute("UPDATE usuarios SET activo = %s WHERE id = %s",
                (0 if row['activo'] else 1, id))
        flash('Usuario actualizado.', 'success')
    return redirect(url_for('gestion_usuarios'))


@app.route('/usuarios/<int:id>/reset-password', methods=['POST'])
def admin_reset_password(id):
    if not current_user.is_admin():
        return redirect(url_for('dashboard'))
    password = request.form.get('password', '')
    if len(password) < 8:
        flash('La contraseña debe tener mínimo 8 caracteres.', 'error')
    else:
        execute("UPDATE usuarios SET password_hash = %s WHERE id = %s",
                (generate_password_hash(password), id))
        flash('Contraseña restablecida.', 'success')
    return redirect(url_for('gestion_usuarios'))


@app.route('/usuarios/<int:id>/eliminar', methods=['POST'])
def eliminar_usuario(id):
    if not current_user.is_admin():
        return redirect(url_for('dashboard'))
    if str(id) == current_user.id:
        flash('No puedes eliminar tu propio usuario.', 'error')
        return redirect(url_for('gestion_usuarios'))
    execute("DELETE FROM usuarios WHERE id = %s", (id,))
    flash('Usuario eliminado.', 'success')
    return redirect(url_for('gestion_usuarios'))


@app.route('/perfil', methods=['GET', 'POST'])
def perfil():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'nombre':
            nombre = request.form.get('nombre', '').strip()
            if nombre:
                execute("UPDATE usuarios SET nombre = %s WHERE id = %s",
                        (nombre, current_user.id))
                flash('Nombre actualizado.', 'success')
        elif action == 'password':
            actual   = request.form.get('actual', '')
            nueva    = request.form.get('nueva', '')
            confirmar = request.form.get('confirmar', '')
            row = fetchone("SELECT password_hash FROM usuarios WHERE id = %s",
                           (current_user.id,))
            if not check_password_hash(row['password_hash'], actual):
                flash('La contraseña actual no es correcta.', 'error')
            elif len(nueva) < 8:
                flash('La nueva contraseña debe tener mínimo 8 caracteres.', 'error')
            elif nueva != confirmar:
                flash('Las contraseñas no coinciden.', 'error')
            else:
                execute("UPDATE usuarios SET password_hash = %s WHERE id = %s",
                        (generate_password_hash(nueva), current_user.id))
                flash('Contraseña actualizada.', 'success')
        return redirect(url_for('perfil'))

    usuario = fetchone("SELECT * FROM usuarios WHERE id = %s", (current_user.id,))
    return render_template('perfil.html', usuario=usuario)


# =============================================================================
#  Soporte Recurrente
# =============================================================================

TIPOS_SOPORTE = ['Sistema', 'SQL', 'Proceso', 'Integración', 'Otro']


def get_temas_soporte(busqueda='', tipo='', division_id=None):
    conditions = ["t.activo = 1"]
    params = []
    if busqueda:
        conditions.append("(LOWER(t.titulo) LIKE %s OR LOWER(t.descripcion) LIKE %s)")
        params += [f'%{busqueda.lower()}%', f'%{busqueda.lower()}%']
    if tipo:
        conditions.append("t.tipo = %s")
        params.append(tipo)
    if division_id:
        conditions.append("t.division_id = %s")
        params.append(division_id)
    where = 'WHERE ' + ' AND '.join(conditions)
    return fetchall(f"""
        SELECT t.*, d.nombre AS division_nombre,
               COUNT(DISTINCT c.id) AS total_causas
        FROM temas_soporte t
        LEFT JOIN divisiones d ON t.division_id = d.id
        LEFT JOIN causas_soporte c ON c.tema_id = t.id
        {where}
        GROUP BY t.id, d.nombre
        ORDER BY t.creado_en DESC
    """, params)


def get_tema_detalle(tema_id):
    tema = fetchone("""
        SELECT t.*, d.nombre AS division_nombre
        FROM temas_soporte t
        LEFT JOIN divisiones d ON t.division_id = d.id
        WHERE t.id = %s
    """, (tema_id,))
    if not tema:
        return None
    causas = fetchall(
        "SELECT * FROM causas_soporte WHERE tema_id = %s ORDER BY id",
        (tema_id,)
    )
    for c in causas:
        c['soluciones'] = fetchall(
            "SELECT * FROM soluciones_soporte WHERE causa_id = %s ORDER BY id",
            (c['id'],)
        )
    tema['causas'] = causas
    tema['entregables'] = enrich_entregables(fetchall("""
        SELECT e.* FROM entregables e
        JOIN tema_entregable te ON te.entregable_id = e.id
        WHERE te.tema_id = %s
        ORDER BY e.id
    """, (tema_id,)))
    return tema


@app.route('/soporte')
def soporte():
    busqueda   = request.args.get('q', '').strip()
    tipo       = request.args.get('tipo', '').strip()
    division_id = request.args.get('division', type=int)
    temas = get_temas_soporte(busqueda, tipo, division_id)
    return render_template('soporte.html',
        temas=temas,
        tipos=TIPOS_SOPORTE,
        divisiones=get_catalog_divisiones(),
        busqueda=busqueda,
        tipo_sel=tipo,
        division_sel=division_id,
    )


@app.route('/soporte/nuevo', methods=['POST'])
def crear_tema_soporte():
    titulo      = request.form.get('titulo', '').strip()
    descripcion = request.form.get('descripcion', '').strip() or None
    tipo        = request.form.get('tipo', 'Sistema')
    url_proyecto = request.form.get('url_proyecto', '').strip() or None
    division_id = request.form.get('division_id') or None
    if not titulo:
        flash('El título es obligatorio.', 'error')
        return redirect(url_for('soporte'))
    execute(
        "INSERT INTO temas_soporte (titulo, descripcion, tipo, url_proyecto, division_id) "
        "VALUES (%s,%s,%s,%s,%s)",
        (titulo, descripcion, tipo, url_proyecto, division_id)
    )
    flash(f'Tema "{titulo}" creado.', 'success')
    return redirect(url_for('soporte'))


@app.route('/soporte/<int:id>')
def detalle_soporte(id):
    tema = get_tema_detalle(id)
    if not tema:
        return 'Tema no encontrado', 404
    all_ents = enrich_entregables(get_entregables())
    vinculados_ids = [e['id'] for e in tema['entregables']]
    disponibles = [e for e in all_ents if e['id'] not in vinculados_ids]
    return render_template('soporte_detalle.html',
        tema=tema,
        disponibles=disponibles,
        divisiones=get_catalog_divisiones(),
        tipos=TIPOS_SOPORTE,
    )


@app.route('/soporte/<int:id>/editar', methods=['POST'])
def editar_tema_soporte(id):
    execute("""
        UPDATE temas_soporte SET titulo=%s, descripcion=%s, tipo=%s,
               url_proyecto=%s, division_id=%s
        WHERE id=%s
    """, (
        request.form.get('titulo', '').strip(),
        request.form.get('descripcion', '').strip() or None,
        request.form.get('tipo', 'Sistema'),
        request.form.get('url_proyecto', '').strip() or None,
        request.form.get('division_id') or None,
        id,
    ))
    flash('Tema actualizado.', 'success')
    return redirect(url_for('detalle_soporte', id=id))


@app.route('/soporte/<int:id>/eliminar', methods=['POST'])
def eliminar_tema_soporte(id):
    execute("DELETE FROM temas_soporte WHERE id = %s", (id,))
    flash('Tema eliminado.', 'success')
    return redirect(url_for('soporte'))


# ── Causas ────────────────────────────────────────────────────────────────

@app.route('/soporte/<int:tema_id>/causas', methods=['POST'])
def agregar_causa(tema_id):
    descripcion = request.form.get('descripcion', '').strip()
    if not descripcion:
        flash('La descripción de la causa es obligatoria.', 'error')
    else:
        execute(
            "INSERT INTO causas_soporte (tema_id, descripcion) VALUES (%s,%s)",
            (tema_id, descripcion)
        )
        flash('Causa agregada.', 'success')
    return redirect(url_for('detalle_soporte', id=tema_id))


@app.route('/soporte/causas/<int:causa_id>/eliminar', methods=['POST'])
def eliminar_causa(causa_id):
    causa = fetchone("SELECT tema_id FROM causas_soporte WHERE id = %s", (causa_id,))
    execute("DELETE FROM causas_soporte WHERE id = %s", (causa_id,))
    flash('Causa eliminada.', 'success')
    return redirect(url_for('detalle_soporte', id=causa['tema_id']) if causa else url_for('soporte'))


# ── Soluciones ────────────────────────────────────────────────────────────

@app.route('/soporte/causas/<int:causa_id>/soluciones', methods=['POST'])
def agregar_solucion(causa_id):
    causa = fetchone("SELECT tema_id FROM causas_soporte WHERE id = %s", (causa_id,))
    descripcion = request.form.get('descripcion', '').strip()
    codigo_sql  = request.form.get('codigo_sql', '').strip() or None
    url_ref     = request.form.get('url_ref', '').strip() or None
    if not descripcion:
        flash('La descripción de la solución es obligatoria.', 'error')
    else:
        execute(
            "INSERT INTO soluciones_soporte (causa_id, descripcion, codigo_sql, url_ref) "
            "VALUES (%s,%s,%s,%s)",
            (causa_id, descripcion, codigo_sql, url_ref)
        )
        flash('Solución agregada.', 'success')
    return redirect(url_for('detalle_soporte', id=causa['tema_id']) if causa else url_for('soporte'))


@app.route('/soporte/soluciones/<int:sol_id>/eliminar', methods=['POST'])
def eliminar_solucion(sol_id):
    sol = fetchone("""
        SELECT c.tema_id FROM soluciones_soporte s
        JOIN causas_soporte c ON s.causa_id = c.id
        WHERE s.id = %s
    """, (sol_id,))
    execute("DELETE FROM soluciones_soporte WHERE id = %s", (sol_id,))
    flash('Solución eliminada.', 'success')
    return redirect(url_for('detalle_soporte', id=sol['tema_id']) if sol else url_for('soporte'))


# ── Entregables vinculados ────────────────────────────────────────────────

@app.route('/soporte/<int:tema_id>/vincular-entregable', methods=['POST'])
def vincular_entregable_soporte(tema_id):
    entregable_id = request.form.get('entregable_id', type=int)
    if entregable_id:
        execute(
            "INSERT INTO tema_entregable (tema_id, entregable_id) VALUES (%s,%s) "
            "ON CONFLICT DO NOTHING",
            (tema_id, entregable_id)
        )
        flash('Entregable vinculado al tema.', 'success')
    return redirect(url_for('detalle_soporte', id=tema_id))


@app.route('/soporte/<int:tema_id>/desvincular-entregable/<int:ent_id>', methods=['POST'])
def desvincular_entregable_soporte(tema_id, ent_id):
    execute(
        "DELETE FROM tema_entregable WHERE tema_id=%s AND entregable_id=%s",
        (tema_id, ent_id)
    )
    flash('Entregable desvinculado.', 'success')
    return redirect(url_for('detalle_soporte', id=tema_id))


# =============================================================================
#  Carga masiva por Layout Excel
# =============================================================================

@app.route('/entregables/descargar-layout')
def descargar_layout():
    """Descarga el layout Excel vacío para carga masiva."""
    path = os.path.join(app.root_path, 'static', 'layout_entregables.xlsx')
    return send_file(path, as_attachment=True, download_name='layout_entregables.xlsx')


@app.route('/entregables/cargar-layout', methods=['POST'])
def cargar_layout():
    """Procesa un layout Excel y crea los entregables."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        flash('Instala openpyxl: pip install openpyxl', 'error')
        return redirect(url_for('entregables'))

    archivo = request.files.get('archivo_layout')
    if not archivo or not archivo.filename.endswith('.xlsx'):
        flash('Selecciona un archivo .xlsx válido.', 'error')
        return redirect(url_for('entregables'))

    # Construir mapas nombre→id para los catálogos
    resp_map  = {r['nombre'].strip().lower(): r['id'] for r in get_catalog_responsables()}
    div_map   = {d['nombre'].strip().lower(): d['id'] for d in get_catalog_divisiones()}
    ku_map    = {k['nombre'].strip().lower(): k['id'] for k in get_catalog_key_users()}
    proy_map  = {p['nombre'].strip().lower(): p['id'] for p in get_catalog_proyectos()}

    wb = load_workbook(filename=io.BytesIO(archivo.read()), data_only=True)
    ws = wb.active  # Primera hoja: Entregables

    ESTATUS_VALIDOS = {'Borrador', 'En progreso', 'En revisión', 'Completado', 'Cancelado'}

    insertados = 0
    errores    = []

    # Las filas de datos comienzan en la 4 (filas 1-3 son título, instrucciones y cabecera)
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Ignorar filas completamente vacías
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        def cel(idx):
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else None

        nombre = cel(1)
        if not nombre:
            errores.append(f'Fila {i}: el campo Nombre es obligatorio.')
            continue

        fecha_entrega_raw = cel(5)

        def parse_fecha(val):
            """Acepta dd/mm/aaaa, aaaa-mm-dd o datetime de Excel."""
            if val is None:
                return None
            if isinstance(val, (datetime, date)):
                return val.date() if isinstance(val, datetime) else val
            val = str(val).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(val, fmt).date()
                except ValueError:
                    continue
            return None

        def yn_to_int(val):
            if val is None:
                return 0
            return 1 if str(val).strip().lower() in ('sí', 'si', '1', 'true', 'sí') else 0

        def safe_int(val, default=None):
            try:
                return int(float(str(val))) if val not in (None, '') else default
            except (ValueError, TypeError):
                return default

        # Columnas: A=0 Folio B=1 Nombre C=2 Desc D=3 FIni E=4 FFin F=5 FEntr G=6 FPub
        #           H=7 Resp I=8 Div J=9 KUser K=10 Proyecto L=11 Horas M=12 Avance
        #           N=13 Estatus O=14 DevOps P=15 FRDAplica Q=16 URLfrd
        resp_nombre = cel(7)
        div_nombre  = cel(8)
        ku_nombre   = cel(9)
        proy_nombre = cel(10)
        estatus_raw = cel(13) or 'Borrador'
        if estatus_raw not in ESTATUS_VALIDOS:
            estatus_raw = 'Borrador'

        division_id = div_map.get(div_nombre.lower()) if div_nombre else None

        payload = {
            'nombre':              nombre,
            'descripcion':         cel(2),
            'fecha_inicio':        parse_fecha(cel(3)),
            'fecha_fin_desarrollo':parse_fecha(cel(4)),
            'fecha_entrega':       parse_fecha(fecha_entrega_raw),
            'fecha_publicacion':   parse_fecha(cel(6)),
            'responsable_id':      resp_map.get(resp_nombre.lower()) if resp_nombre else None,
            'division_id':         division_id,
            'key_user_id':         ku_map.get(ku_nombre.lower())     if ku_nombre   else None,
            'proyecto_id':         proy_map.get(proy_nombre.lower()) if proy_nombre else None,
            'horas_desarrollo':    safe_int(cel(11)),
            'porcentaje_avance':   safe_int(cel(12), 0),
            'estatus':             estatus_raw,
            'necesita_devops':     yn_to_int(cel(14)),
            'frd_aplica':          yn_to_int(cel(15)),
            'frd_url':             cel(16),
            'frd_firmado':         0,
            'frd_aceptacion':      0,
            'gherkin':             None,
            'creado_en':           datetime.now(),
        }
        # Folio automático basado en la división (igual que la captura manual)
        payload['folio'] = generate_folio(division_id)

        if fecha_entrega_raw and payload['fecha_entrega'] is None:
            errores.append(f'Fila {i}: Fecha Entrega "{fecha_entrega_raw}" no tiene formato válido (dd/mm/aaaa).')
            continue

        try:
            insert_entregable(payload)
            insertados += 1
        except Exception as e:
            errores.append(f'Fila {i}: error al insertar — {e}')

    if insertados:
        flash(f'✅ {insertados} entregable{"s" if insertados != 1 else ""} importado{"s" if insertados != 1 else ""} correctamente.', 'success')
    if errores:
        flash('⚠️ Se encontraron errores en algunas filas: ' + ' | '.join(errores[:5])
              + (f' … y {len(errores)-5} más.' if len(errores) > 5 else ''), 'warning')
    if not insertados and not errores:
        flash('El archivo no contiene filas de datos.', 'warning')

    return redirect(url_for('entregables'))


if __name__ == '__main__':
    if not use_pg:
        init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
