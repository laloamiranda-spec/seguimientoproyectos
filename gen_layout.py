"""
Genera el layout Excel para carga masiva de entregables.
Ejecutar con: py gen_layout.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT_PATH = r"C:\Desarrollo\Entregables\static\layout_entregables.xlsx"

# ---------------------------------------------------------------------------
# Colores (ARGB)
# ---------------------------------------------------------------------------
C_INDIGO       = "FF4F46E5"   # #4F46E5 — título / headers catálogos
C_INDIGO_LIGHT = "FF6366F1"   # #6366F1 — headers columnas
C_INDIGO_BG    = "FFEEF2FF"   # #EEF2FF — instrucciones
C_DATA_ALT     = "FFF8F9FF"   # #F8F9FF — filas alternas datos
C_WHITE        = "FFFFFFFF"
C_GRAY_TEXT    = "FF374151"   # #374151 — texto instrucciones
C_GRAY_LIGHT   = "FFF3F4F6"   # #F3F4F6 — catálogos bg
C_BORDER       = "FFD1D5DB"   # gris claro bordes

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def thin_border(color=C_BORDER):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def fill(argb):
    return PatternFill("solid", fgColor=argb)

def font(name="Arial", size=10, bold=False, italic=False, color="FF000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def align(horizontal="left", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# ============================================================
# SHEET 1 — Entregables
# ============================================================
ws = wb.active
ws.title = "Entregables"

# --- Column definitions ---
columns = [
    ("A", "Folio",                  12),
    ("B", "Nombre *",               30),   # required → suffix already in header
    ("C", "Descripción",            35),
    ("D", "Fecha Inicio",           16),
    ("E", "Fecha Fin Desarrollo",   20),
    ("F", "Fecha Entrega *",        16),   # required
    ("G", "Fecha Publicación",      18),
    ("H", "Responsable",            22),
    ("I", "División",               22),
    ("J", "Key User",               22),
    ("K", "Horas Desarrollo",       16),
    ("L", "% Avance",               10),
    ("M", "Estatus",                18),
    ("N", "Necesita DevOps",        14),
    ("O", "FRD Aplica",             10),
    ("P", "URL FRD",                35),
]

# Set column widths
for col_letter, _, width in columns:
    ws.column_dimensions[col_letter].width = width

# Row heights
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 22

# ── ROW 1: Title ────────────────────────────────────────────
ws.merge_cells("A1:P1")
title_cell = ws["A1"]
title_cell.value = "LAYOUT CARGA DE ENTREGABLES"
title_cell.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
title_cell.fill = fill(C_INDIGO)
title_cell.alignment = align("center", "center")

# ── ROW 2: Instructions ──────────────────────────────────────
ws.merge_cells("A2:P2")
instr_cell = ws["A2"]
instr_cell.value = (
    "* Campos obligatorios: Nombre, Fecha Entrega. "
    "Usa los catálogos de la hoja 'Catálogos' para Responsable, División y Key User."
)
instr_cell.font = Font(name="Arial", size=9, italic=True, color=C_GRAY_TEXT)
instr_cell.fill = fill(C_INDIGO_BG)
instr_cell.alignment = align("left", "center", wrap=False)

# ── ROW 3: Column headers ────────────────────────────────────
header_fill = fill(C_INDIGO_LIGHT)
header_font = Font(name="Arial", size=10, bold=True, color=C_WHITE)
header_border = thin_border("FF4338CA")

for col_letter, header_text, _ in columns:
    cell = ws[f"{col_letter}3"]
    cell.value = header_text
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align("center", "center")
    cell.border = header_border

# ── ROWS 4-6: Example data ───────────────────────────────────
example_rows = [
    [
        "ENT-001", "Portal de clientes", "Módulo de login y dashboard",
        "01/01/2026", "15/03/2026", "31/03/2026", "05/04/2026",
        "", "", "",
        40, 65, "En progreso", "No", "Sí", "https://drive.google.com/doc1"
    ],
    [
        "ENT-002", "Reporte mensual ventas", "Automatización de reporte Excel",
        "10/01/2026", "28/02/2026", "15/03/2026", "",
        "", "", "",
        16, 100, "Completado", "No", "No", ""
    ],
    [
        "ENT-003", "Integración ERP-CRM", "Sincronización de pedidos",
        "01/02/2026", "30/04/2026", "30/04/2026", "",
        "", "", "",
        80, 10, "Borrador", "Sí", "Sí", "https://drive.google.com/doc3"
    ],
]

data_font   = Font(name="Arial", size=9, color="FF1F2937")
data_border = thin_border(C_BORDER)
data_fill   = fill(C_DATA_ALT)

for row_idx, row_data in enumerate(example_rows, start=4):
    ws.row_dimensions[row_idx].height = 16
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = align("left", "center")

# ── FREEZE pane at row 4 (headers visible) ───────────────────
ws.freeze_panes = "A4"

# ── DATA VALIDATIONS (rows 4:500) ───────────────────────────
# Estatus — column M
dv_estatus = DataValidation(
    type="list",
    formula1='"Borrador,En progreso,En revisión,Completado,Cancelado"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Valor inválido",
    error="Selecciona un estatus de la lista."
)
dv_estatus.sqref = "M4:M500"
ws.add_data_validation(dv_estatus)

# Necesita DevOps — column N
dv_devops = DataValidation(
    type="list",
    formula1='"Sí,No"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Valor inválido",
    error="Ingresa 'Sí' o 'No'."
)
dv_devops.sqref = "N4:N500"
ws.add_data_validation(dv_devops)

# FRD Aplica — column O
dv_frd = DataValidation(
    type="list",
    formula1='"Sí,No"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Valor inválido",
    error="Ingresa 'Sí' o 'No'."
)
dv_frd.sqref = "O4:O500"
ws.add_data_validation(dv_frd)

# ── Tab color ────────────────────────────────────────────────
ws.sheet_properties.tabColor = "4F46E5"


# ============================================================
# SHEET 2 — Catálogos
# ============================================================
ws2 = wb.create_sheet("Catálogos")
ws2.sheet_properties.tabColor = "6366F1"

cat_headers = ["Responsables", "Divisiones", "Key Users"]
cat_cols    = ["A", "B", "C"]

# Column widths
for col_letter in cat_cols:
    ws2.column_dimensions[col_letter].width = 30

ws2.row_dimensions[1].height = 22
ws2.row_dimensions[2].height = 40

cat_header_font   = Font(name="Arial", size=10, bold=True, color=C_WHITE)
cat_header_fill   = fill(C_INDIGO)
cat_header_border = thin_border("FF4338CA")

cat_body_font   = Font(name="Arial", size=9, italic=True, color="FF6B7280")
cat_body_fill   = fill(C_GRAY_LIGHT)
cat_body_border = thin_border(C_BORDER)

for col_letter, header_text in zip(cat_cols, cat_headers):
    # Header
    hcell = ws2[f"{col_letter}1"]
    hcell.value = header_text
    hcell.font = cat_header_font
    hcell.fill = cat_header_fill
    hcell.alignment = align("center", "center")
    hcell.border = cat_header_border

    # Body
    bcell = ws2[f"{col_letter}2"]
    bcell.value = "Consulta la app para ver el catálogo actual"
    bcell.font = cat_body_font
    bcell.fill = cat_body_fill
    bcell.alignment = align("center", "center", wrap=True)
    bcell.border = cat_body_border

# Light gray fill on remaining rows 3-50 to reinforce read-only appearance
for row in range(3, 51):
    for col_letter in cat_cols:
        cell = ws2[f"{col_letter}{row}"]
        cell.fill = fill(C_GRAY_LIGHT)
        cell.border = thin_border(C_BORDER)

# ── Freeze header row ────────────────────────────────────────
ws2.freeze_panes = "A2"

# ============================================================
# Save
# ============================================================
wb.save(OUTPUT_PATH)
print(f"Archivo generado exitosamente: {OUTPUT_PATH}")
